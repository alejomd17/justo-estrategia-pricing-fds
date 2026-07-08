"""Estrategia de ofertas FDS: elegibilidad, score de priorizacion, cascada
de mecanica, guardrail economico, dia de ejecucion y exportacion a Excel.

Extraido de `src/exploration.ipynb` para que tanto el notebook como (en el
futuro) el backend puedan reusar la misma logica sin duplicarla.
"""

import re
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from . import catalog, elasticity, rotacion

# ---------- Reglas de margen ----------
PISO_MARGEN = 0.22
TOPE_GUARDRAIL = 0.30  # autonomo; por encima requiere Vo.Bo. de Comercial
DESC_MINIMO = 0.05  # por debajo de esto no vale la pena publicar la oferta

# ---------- Priorizacion (score -> fraccion del colchon autorizada) ----------
TOPES_SCORE = [(7, 0.25), (5, 0.20), (3, 0.15), (2, 0.10)]
CASTIGO_MAS_BARATO = 0.05  # se resta al tope si el SKU ya es "Mas barato" en linea
BOOST_MUNDIAL = 1
TOLERANCIA_CASCADA = 0.02

# ---------- Mecanicas flexibles por ticket ----------
TICKET_ALTO = 150  # BNSDP: ahorro en pesos, ticket alto
TICKET_BAJO = 60  # BNSP: paquete de 3 a precio redondo, alta rotacion
UMBRALES_MULTIBUY = [(0.25, "4x3", 3, 4), (0.20, "5x4", 4, 5), (1 / 6, "6x5", 5, 6)]

# ---------- Modelo economico ----------
MULT_PROMO = 2.0  # amplificacion promocional declarada; se contrasta contra
# el backtesting para ver si el supuesto se sostiene
REDENCION = {
    "multibuy_grande": 0.35,  # 4x3 / 5x4 / 6x5
    "pack_3": 0.40,  # BNSP 3 unidades
    "umbral_2": 0.55,  # SPON / BNSDP 2 unidades
}
ESCENARIOS = {"Conservador": 0.6, "Base": 1.0, "Optimista": 1.4}

# ---------- Tema Mundial ----------
PATRON_MUNDIAL = re.compile(
    r"\b(?:"
    r"cerveza|refresco|"
    r"papas?\s+(?:fritas?|pringles|ruffles)|totopos?|nachos?|chicharr\w*|"
    r"palomitas|botanas?|"
    r"habanero|valentina|cholula|clamato|b[uú]falo"
    r")\b",
    re.IGNORECASE,
)

# ---------- Pesos del score ----------
ELAS_W = {"ALTAMENTE ELASTICO": 3, "ELASTICO": 2, "POCO ELASTICO": 1, "INELASTICO": 0, "SIN DATOS": 0}
ROT_W = {"Alta rotacion": 3, "Rotacion normal": 2, "Baja rotacion": 1, "Sin rotacion": 0}
# Bono real (no el TAG_ROTACION estatico) por rotar mas rapido que el
# promedio de su categoria (rotacion.py) - acotado a [-1, +1], mismo orden
# de magnitud que el bono de TAG_SHARE/Mundial, para no dominar el score.
PESO_ROTACION_REAL = 1.0


def cargar_oportunidad(ruta_oportunidad: str) -> pd.DataFrame:
    """Carga y limpia el universo: quita filas de resumen/filtros del export
    y separa SKU/Nombre."""
    oportunidad = pd.read_excel(ruta_oportunidad, sheet_name="Export")
    df = oportunidad[oportunidad["SKU + Nombre"].notna() & oportunidad["COSTO"].notna()].copy()
    df["STORE_ID"] = pd.to_numeric(df["STORE_ID"], errors="coerce")
    df = df[df["STORE_ID"].notna()].copy()
    df["STORE_ID"] = df["STORE_ID"].astype(int)

    df[["SKU", "Nombre"]] = df["SKU + Nombre"].str.split(" - ", n=1, expand=True)
    df["SKU"] = pd.to_numeric(df["SKU"], errors="coerce")
    df = df.dropna(subset=["SKU"]).copy()
    df["SKU"] = df["SKU"].astype(int)

    df["MOTIVO_SIN_OFERTA"] = ""
    return df


def _pares_limpios(tabla: pd.DataFrame, cols=("ItemId", "AreaId")) -> set:
    t = tabla[list(cols)].apply(pd.to_numeric, errors="coerce").dropna().astype(int)
    return set(map(tuple, t.values))


def excluir_comercial(
    df: pd.DataFrame, ruta_descuentos: str, weekend_inicio: pd.Timestamp, weekend_fin: pd.Timestamp
) -> pd.DataFrame:
    """Marca (no borra) las filas en Black list o con promo comercial
    vigente que cruce el FDS. Une `DEscuentos comerciales` + `Historico`
    (Estatus == 'Cargado'), y resta `Eliminacion de campañas`."""
    black_list = pd.read_excel(ruta_descuentos, sheet_name="Black list")
    pares_blacklist = _pares_limpios(black_list)

    comerciales = pd.read_excel(ruta_descuentos, sheet_name="DEscuentos comerciales")
    historico = pd.read_excel(ruta_descuentos, sheet_name="Historico")
    historico = historico[historico["Estatus"] == "Cargado"]

    promos = pd.concat([comerciales, historico], ignore_index=True)
    promos["ItemId"] = pd.to_numeric(promos["ItemId"], errors="coerce")
    promos["AreaId"] = pd.to_numeric(promos["AreaId"], errors="coerce")
    promos = promos.dropna(subset=["ItemId", "AreaId", "FechaInicio", "FechaFin"]).copy()
    promos["ItemId"] = promos["ItemId"].astype(int)
    promos["AreaId"] = promos["AreaId"].astype(int)

    cruza_fds = promos[
        (promos["FechaInicio"] <= weekend_fin) & (promos["FechaFin"] >= weekend_inicio)
    ]
    pares_promo = set(zip(cruza_fds["ItemId"], cruza_fds["AreaId"]))

    eliminadas = pd.read_excel(ruta_descuentos, sheet_name="Eliminacion de campañas")
    pares_eliminados = _pares_limpios(eliminadas)
    pares_promo_vigente = pares_promo - pares_eliminados

    df_pares = pd.Series(list(zip(df["SKU"], df["STORE_ID"])), index=df.index)
    en_blacklist = df_pares.isin(pares_blacklist)
    en_promo = df_pares.isin(pares_promo_vigente) & ~en_blacklist

    df.loc[en_blacklist, "MOTIVO_SIN_OFERTA"] = "Black list"
    df.loc[en_promo, "MOTIVO_SIN_OFERTA"] = "Promo comercial vigente en el FDS"
    return df


def agregar_catalogo_real(df: pd.DataFrame, cur) -> pd.DataFrame:
    """Merge con Departamento/Categoria real (FULL_MASTER_CATALOG)."""
    catalogo_df = catalog.get_full_master_catalog(cur)
    return df.merge(catalogo_df, on="SKU", how="left")


def agregar_rotacion_real(df: pd.DataFrame, cur) -> pd.DataFrame:
    """Merge con el baseline real de rotacion por categoria (rotacion.py,
    catalogo completo - no acotado al universo de oportunidad.xlsx) y
    calcula ROTACION_REAL_RATIO = rotacion propia del SKU / baseline de su
    categoria. df ya debe traer Categoria (merge previo con
    agregar_catalogo_real)."""
    baseline = rotacion.get_baseline_categoria(cur)
    cols_baseline = [
        "SKU",
        "STORE_ID",
        "UNIDADES_DIA_SKU_TIENDA",
        "DIAS_ENTRE_VENTA_SKU_TIENDA",
        "UNIDADES_DIA_CATEGORIA_BASELINE",
        "DIAS_ENTRE_VENTA_CATEGORIA_BASELINE",
        "FUENTE_BASELINE_CATEGORIA",
    ]
    df = df.merge(baseline[cols_baseline], on=["SKU", "STORE_ID"], how="left")
    df["ROTACION_REAL_RATIO"] = df["UNIDADES_DIA_SKU_TIENDA"] / df["UNIDADES_DIA_CATEGORIA_BASELINE"]
    return df


def calcular_margen_neto(df: pd.DataFrame) -> pd.DataFrame:
    """Agrega PRECIO_NETO (util para pesos de utilidad mas abajo). No
    reconstruye MARGEN - esa columna del export es la fuente de verdad."""
    df["PRECIO_NETO"] = (
        df["PRECIO JUSTO"] / (1 + df["IVA"].fillna(0) / 100) / (1 + df["IEPS"].fillna(0) / 100)
    )
    return df


def marcar_elegibilidad(df: pd.DataFrame) -> pd.DataFrame:
    """Margen base < 22% o TAG ELAS inelastico/sin datos -> sin oferta."""
    sin_motivo = df["MOTIVO_SIN_OFERTA"] == ""

    m_bajo = (df["MARGEN"] < PISO_MARGEN * 100) & sin_motivo
    df.loc[m_bajo, "MOTIVO_SIN_OFERTA"] = "Margen base < 22%"

    elas_out = df["TAG ELAS"].isin(["INELASTICO", "SIN DATOS"]) & sin_motivo & ~m_bajo
    df.loc[elas_out, "MOTIVO_SIN_OFERTA"] = "Inelastico / sin datos de elasticidad"

    df["ELEGIBLE"] = df["MOTIVO_SIN_OFERTA"] == ""
    return df


def detectar_tema_mundial(df: pd.DataFrame) -> pd.DataFrame:
    df["TEMA_MUNDIAL"] = df["Nombre"].str.contains(PATRON_MUNDIAL, na=False)
    return df


def calcular_dmax(df: pd.DataFrame) -> pd.DataFrame:
    """DMAX_REAL: colchon puro de margen. DMAX: acotado al tope autonomo de
    30%. REQUIERE_APROBACION: True si el colchon real supera ese tope."""
    m = df["MARGEN"] / 100
    df["DMAX_REAL"] = (1 - (1 - m) / (1 - PISO_MARGEN)).clip(lower=0)
    df.loc[~df["ELEGIBLE"], "DMAX_REAL"] = 0.0

    df["DMAX"] = df["DMAX_REAL"].clip(upper=TOPE_GUARDRAIL)
    df["REQUIERE_APROBACION"] = df["DMAX_REAL"] > TOPE_GUARDRAIL
    return df


def _tope_por_score(s):
    for smin, tope in TOPES_SCORE:
        if s >= smin:
            return tope
    return 0.0


def calcular_score_y_tope(df: pd.DataFrame) -> pd.DataFrame:
    df["W_ELAS"] = df["TAG ELAS"].map(ELAS_W).fillna(0)
    df["W_ROT"] = df["TAG_ROTACION"].map(ROT_W).fillna(0)
    # Ratio real de rotacion vs. su categoria (rotacion=1 -> en linea con la
    # categoria) llevado a un bono acotado [-1, +1] - sin baseline (NaN) no
    # suma ni resta.
    bono_rotacion_real = (df["ROTACION_REAL_RATIO"].clip(upper=2) - 1).clip(lower=-1).fillna(0)
    df["SCORE"] = (
        df["W_ELAS"] * df["W_ROT"]
        + (df["TAG_SHARE"] == "ALTO SHARE").astype(int)
        + df["TEMA_MUNDIAL"].astype(int) * BOOST_MUNDIAL
        + bono_rotacion_real * PESO_ROTACION_REAL
    )

    df["TOPE"] = df["SCORE"].apply(_tope_por_score)
    mas_barato = df["INDEX PRECIO LINEA"] == "Mas barato"
    df.loc[mas_barato, "TOPE"] = (df.loc[mas_barato, "TOPE"] - CASTIGO_MAS_BARATO).clip(lower=0)

    df["D_TARGET"] = np.minimum(df["DMAX"], df["TOPE"])
    df.loc[~df["ELEGIBLE"], "D_TARGET"] = 0.0
    return df


def _asignar_mecanica(r):
    d, dmax, p = r["D_TARGET"], r["DMAX"], r["PRECIO JUSTO"]
    if d < DESC_MINIMO:
        return pd.Series(["Sin oferta", p, 0.0])
    for umbral, nombre, n_pag, n_llev in UMBRALES_MULTIBUY:
        if d >= umbral or ((umbral - d <= TOLERANCIA_CASCADA) and (dmax >= umbral)):
            return pd.Series([nombre, round(p * n_pag / n_llev, 2), round(umbral, 4)])
    d5 = np.floor(d * 20) / 20
    if d5 < DESC_MINIMO:
        return pd.Series(["Sin oferta", p, 0.0])
    if p >= TICKET_ALTO:
        ahorro = np.floor(d5 * p / 5) * 5
        if ahorro < 5:
            return pd.Series(["Sin oferta", p, 0.0])
        return pd.Series([f"BNSDP: 2 uds, ahorra ${ahorro:.0f} c/u", round(p - ahorro, 2), round(ahorro / p, 4)])
    if p <= TICKET_BAJO and r["W_ROT"] == 3:
        pack = np.floor(3 * p * (1 - d5))
        return pd.Series([f"BNSP: 3 x ${pack:.0f}", round(pack / 3, 2), round(1 - pack / (3 * p), 4)])
    return pd.Series([f"SPON: 2 uds, -{d5*100:.0f}%", round(p * (1 - d5), 2), d5])


def asignar_mecanica(df: pd.DataFrame) -> pd.DataFrame:
    df[["ESTRATEGIA", "PRECIO_OFERTA", "DESC_EFECTIVO"]] = df.apply(_asignar_mecanica, axis=1)
    df["MECANICA"] = df["ESTRATEGIA"].str.split(":").str[0]
    return df


def validar_piso(df: pd.DataFrame) -> pd.DataFrame:
    """Recalcula el margen con la mecanica activa; degrada a 'Sin oferta'
    si por redondeo/tolerancia se rompe el piso de 22%."""
    m = df["MARGEN"] / 100
    df["MARGEN_OFERTA"] = (1 - (1 - m) / (1 - df["DESC_EFECTIVO"])) * 100

    viola = (df["MECANICA"] != "Sin oferta") & (df["MARGEN_OFERTA"] < PISO_MARGEN * 100)
    df.loc[viola, ["ESTRATEGIA", "DESC_EFECTIVO"]] = ["Sin oferta", 0.0]
    df.loc[viola, "PRECIO_OFERTA"] = df.loc[viola, "PRECIO JUSTO"]
    df.loc[viola, "MOTIVO_SIN_OFERTA"] = "Degradada: violaria el piso de margen"
    df["MECANICA"] = df["ESTRATEGIA"].str.split(":").str[0]
    return df


def _redencion(mec):
    if mec in ("4x3", "5x4", "6x5"):
        return REDENCION["multibuy_grande"]
    if mec == "BNSP":
        return REDENCION["pack_3"]
    if mec in ("SPON", "BNSDP"):
        return REDENCION["umbral_2"]
    return 0.0


def calcular_modelo_economico(df: pd.DataFrame) -> pd.DataFrame:
    """u = |E| x MULT x d ; Q1 = Q0(1+u). El costo de la promo solo lo paga
    la fraccion redimida r: GMV/utilidad mezclados por (1 - r*d)."""
    m = df["MARGEN"] / 100
    df["REDENCION"] = df["MECANICA"].apply(_redencion)
    d_eff = df["DESC_EFECTIVO"]
    df["UPLIFT"] = df["ELASTICIDAD_FINAL"].abs() * MULT_PROMO * d_eff
    df["Q1_DIA"] = df["Q0_DIA"] * (1 + df["UPLIFT"])

    rd = df["REDENCION"] * d_eff
    df["GMV_BASE_DIA"] = (df["Q0_DIA"] * df["PRECIO JUSTO"]).round(2)
    df["GMV_PROY_DIA"] = (df["Q1_DIA"] * df["PRECIO JUSTO"] * (1 - rd)).round(2)
    df["UTIL_BASE_DIA"] = (df["Q0_DIA"] * df["PRECIO_NETO"] * m).round(2)
    df["UTIL_PROY_DIA"] = (df["Q1_DIA"] * df["PRECIO_NETO"] * (m - rd)).round(2)
    df["GMV_INC_DIA"] = (df["GMV_PROY_DIA"] - df["GMV_BASE_DIA"]).round(2)
    df["UTIL_INC_DIA"] = (df["UTIL_PROY_DIA"] - df["UTIL_BASE_DIA"]).round(2)
    return df


def aplicar_guardrail_economico(df: pd.DataFrame) -> pd.DataFrame:
    """No se publica una oferta si su utilidad incremental proyectada es
    negativa - no tiene sentido descontar algo que ya se vende bien solo."""
    m = df["MARGEN"] / 100
    destruye = (df["MECANICA"] != "Sin oferta") & (df["UTIL_INC_DIA"] < 0)
    df.loc[destruye, ["ESTRATEGIA", "DESC_EFECTIVO"]] = ["Sin oferta", 0.0]
    df.loc[destruye, "PRECIO_OFERTA"] = df.loc[destruye, "PRECIO JUSTO"]
    df.loc[destruye, "MOTIVO_SIN_OFERTA"] = "Descartada: utilidad incremental proyectada < 0"
    df["MECANICA"] = df["ESTRATEGIA"].str.split(":").str[0]

    anular = df["MECANICA"] == "Sin oferta"
    df.loc[anular, ["UPLIFT", "REDENCION"]] = 0.0
    df.loc[anular, "Q1_DIA"] = df.loc[anular, "Q0_DIA"]
    df.loc[anular, "GMV_PROY_DIA"] = df.loc[anular, "GMV_BASE_DIA"]
    df.loc[anular, "UTIL_PROY_DIA"] = df.loc[anular, "UTIL_BASE_DIA"]
    df.loc[anular, ["GMV_INC_DIA", "UTIL_INC_DIA"]] = 0.0
    df["MARGEN_OFERTA"] = ((1 - (1 - m) / (1 - df["DESC_EFECTIVO"])) * 100).round(2)
    return df


def calcular_confianza(df: pd.DataFrame, skus_con_historial_promo: set) -> pd.DataFrame:
    def nivel(row):
        if row["FUENTE_ELASTICIDAD"] != "SKU real":
            return "Baja"
        return "Alta" if (row["SKU"], row["STORE_ID"]) in skus_con_historial_promo else "Media"

    df["CONFIANZA_PROYECCION"] = df.apply(nivel, axis=1)
    return df


def get_promo_historicas(cur, skus_elegibles: set) -> tuple[pd.DataFrame, set]:
    """Historial de promos (DISCOUNT_CAMPAIGN_EVENT) desde 2025-01-01,
    acotado a los SKUs elegibles. Devuelve (promos_relevantes, pares
    (SKU,STORE_ID) con al menos una promo pasada) para Confianza/backtest."""
    cur.execute("""
        SELECT
            REGEXP_REPLACE(PRODUCT_ID, '[^0-9]', '') AS SKU,
            WAREHOUSE_ID                              AS STORE_ID,
            START_DATE,
            END_DATE,
            BY_BULK_STRATEGY_DISCOUNT_TYPE            AS TIPO_MECANICA,
            BY_BULK_STRATEGY_VALUE                    AS VALOR_DESCUENTO
        FROM MX_JUSTO_PROD.ODS_MS_PRICING_V.DISCOUNT_CAMPAIGN_EVENT
        WHERE START_DATE >= '2025-01-01'
    """)
    columnas = [c[0] for c in cur.description]
    promos_historicas = pd.DataFrame(cur.fetchall(), columns=columnas)
    promos_historicas["SKU"] = pd.to_numeric(promos_historicas["SKU"], errors="coerce")
    promos_historicas = promos_historicas.dropna(subset=["SKU", "STORE_ID", "START_DATE", "END_DATE"])
    promos_historicas["SKU"] = promos_historicas["SKU"].astype(int)
    promos_historicas["STORE_ID"] = promos_historicas["STORE_ID"].astype(int)

    promos_relevantes = promos_historicas[promos_historicas["SKU"].isin(skus_elegibles)].copy()
    skus_con_historial_promo = set(zip(promos_relevantes["SKU"], promos_relevantes["STORE_ID"]))
    return promos_relevantes, skus_con_historial_promo


def get_ventas_historicas(cur, skus_elegibles: set) -> pd.DataFrame:
    """Ventas diarias reales (MASTER_ORDERLINE) de 2025-01-01 a 2026-10-22
    para los SKUs elegibles - usadas en el backtesting."""
    skus_param = tuple(int(s) for s in skus_elegibles) if skus_elegibles else (0,)
    placeholders = ",".join(["%s"] * len(skus_param))

    query_ventas = f"""
        SELECT
            PRODUCT_ID::VARCHAR                                    AS SKU,
            STORE_ID,
            TO_DATE(DATETIME_DELIVERY)                             AS FECHA,
            SUM(QUANTITY_FULFILLED_PZ)                             AS UNIDADES
        FROM MX_JUSTO_PROD.DR_MASTER_TABLES.MASTER_ORDERLINE
        WHERE STATUS_ORDER            = 'delivered'
          AND STORE_ID                IN (9, 14)
          AND QUANTITY_FULFILLED_PZ   > 0
          AND DATETIME_DELIVERY       BETWEEN '2025-01-01' AND '2026-10-22'
          AND PRODUCT_ID::VARCHAR     IN ({placeholders})
        GROUP BY 1, 2, 3
        LIMIT 1000000
    """
    cur.execute(query_ventas, skus_param)
    columnas = [c[0] for c in cur.description]
    ventas_historicas = pd.DataFrame(cur.fetchall(), columns=columnas)
    ventas_historicas["SKU"] = pd.to_numeric(ventas_historicas["SKU"], errors="coerce")
    ventas_historicas = ventas_historicas.dropna(subset=["SKU"])
    ventas_historicas["SKU"] = ventas_historicas["SKU"].astype(int)
    ventas_historicas["STORE_ID"] = ventas_historicas["STORE_ID"].astype(int)
    ventas_historicas["FECHA"] = pd.to_datetime(ventas_historicas["FECHA"])
    return ventas_historicas


def calcular_backtesting(
    df: pd.DataFrame, promos_relevantes: pd.DataFrame, ventas_historicas: pd.DataFrame
) -> tuple[pd.DataFrame, float, float]:
    """Compara el uplift que el modelo (con MULT_PROMO actual) habria
    proyectado en promos % historicas contra el uplift real observado.
    Devuelve (backtest_df, mae, bias) - mae/bias son None si no hay datos."""
    promos_pct = promos_relevantes[
        (promos_relevantes["TIPO_MECANICA"] == "PERCENTAGE") & promos_relevantes["VALOR_DESCUENTO"].notna()
    ]
    elasticidad_real_dict = df.set_index(["SKU", "STORE_ID"])["ELASTICIDAD_FINAL"].to_dict()
    fuente_dict = df.set_index(["SKU", "STORE_ID"])["FUENTE_ELASTICIDAD"].to_dict()

    filas_backtest = []
    for _, promo in promos_pct.iterrows():
        sku, store = promo["SKU"], promo["STORE_ID"]
        if fuente_dict.get((sku, store)) != "SKU real":
            continue
        elasticidad = elasticidad_real_dict.get((sku, store))
        if elasticidad is None or pd.isna(elasticidad):
            continue

        ventas_sku = ventas_historicas[
            (ventas_historicas["SKU"] == sku) & (ventas_historicas["STORE_ID"] == store)
        ]
        if ventas_sku.empty:
            continue
        en_promo = ventas_sku[
            (ventas_sku["FECHA"] >= promo["START_DATE"]) & (ventas_sku["FECHA"] <= promo["END_DATE"])
        ]
        fuera_promo = ventas_sku[
            (ventas_sku["FECHA"] < promo["START_DATE"]) | (ventas_sku["FECHA"] > promo["END_DATE"])
        ]
        if len(en_promo) < 2 or len(fuera_promo) < 5:
            continue
        unidades_promo = en_promo["UNIDADES"].mean()
        unidades_base = fuera_promo["UNIDADES"].mean()
        if unidades_base <= 0:
            continue

        d_hist = float(promo["VALOR_DESCUENTO"]) / 100
        uplift_real = (unidades_promo - unidades_base) / unidades_base * 100
        uplift_proyectado = abs(elasticidad) * MULT_PROMO * d_hist * 100

        filas_backtest.append(
            {
                "SKU": sku,
                "STORE_ID": store,
                "Uplift real %": round(uplift_real, 1),
                "Uplift proyectado %": round(uplift_proyectado, 1),
                "Error (proyectado - real)": round(uplift_proyectado - uplift_real, 1),
            }
        )

    backtest = pd.DataFrame(filas_backtest)
    if len(backtest):
        mae = backtest["Error (proyectado - real)"].abs().mean()
        bias = backtest["Error (proyectado - real)"].mean()
    else:
        mae = bias = None
    return backtest, mae, bias


def asignar_dia_ejecucion(df: pd.DataFrame) -> pd.DataFrame:
    """Precedencia: Mundial -> Sab+Dom ; Despensa -> Domingo ; Alta rotacion
    -> Viernes ; BNSDP (ticket alto) -> Domingo ; resto -> Sabado."""

    def dia(r):
        if r["MECANICA"] == "Sin oferta":
            return ""
        if r["TEMA_MUNDIAL"]:
            return "Sabado y Domingo"
        if r["Departamento"] == "Despensa":
            return "Domingo"
        if r["TAG_ROTACION"] == "Alta rotacion":
            return "Viernes"
        if r["MECANICA"] == "BNSDP":
            return "Domingo"
        return "Sabado"

    df["DIA_EJECUCION"] = df.apply(dia, axis=1)
    return df


def calcular_escenarios(df: pd.DataFrame) -> pd.DataFrame:
    filas = []
    o = df[df["MECANICA"] != "Sin oferta"]
    rd_o = o["REDENCION"] * o["DESC_EFECTIVO"]
    for nombre, f in ESCENARIOS.items():
        q1 = o["Q0_DIA"] * (1 + o["UPLIFT"] * f)
        gmv1 = (q1 * o["PRECIO JUSTO"] * (1 - rd_o)).sum()
        util1 = (q1 * o["PRECIO_NETO"] * (o["MARGEN"] / 100 - rd_o)).sum()
        filas.append(
            {
                "Escenario": nombre,
                "Factor uplift": f,
                "Unidades": round(q1.sum()),
                "GMV dia oferta": round(gmv1),
                "GMV incremental": round(gmv1 - o["GMV_BASE_DIA"].sum()),
                "Utilidad incremental": round(util1 - o["UTIL_BASE_DIA"].sum()),
            }
        )
    return pd.DataFrame(filas)


def exportar_excel(
    df: pd.DataFrame,
    backtest: pd.DataFrame,
    escenarios_df: pd.DataFrame,
    ruta_salida: str,
    weekend_inicio: pd.Timestamp,
    weekend_fin: pd.Timestamp,
    mae=None,
    bias=None,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Exporta el Excel final (6 hojas) y devuelve (out_oferta, out_descartados)."""
    ofer = df[df["MECANICA"] != "Sin oferta"]

    cols_out = [
        "STORE_ID", "SKU", "Nombre", "Departamento", "Categoria", "COSTO", "IEPS", "IVA",
        "MARGEN", "PRECIO JUSTO", "TAG_MARGEN", "INDEX PRECIO DCTO", "PRECIO COMP DCTO",
        "PRECIO COMP LINEA", "INDEX PRECIO LINEA", "TAG_ROTACION", "TAG_SHARE", "TAG ELAS",
        "ROTACION_REAL_RATIO", "DIAS_ENTRE_VENTA_SKU_TIENDA", "DIAS_ENTRE_VENTA_CATEGORIA_BASELINE",
        "FUENTE_BASELINE_CATEGORIA",
        "PROMO ACTIVA HOY", "TEMA_MUNDIAL", "MOTIVO_SIN_OFERTA", "REQUIERE_APROBACION",
        "DMAX_REAL", "ESTRATEGIA", "DIA_EJECUCION", "PRECIO_OFERTA", "DESC_EFECTIVO",
        "MARGEN_OFERTA", "FUENTE_ELASTICIDAD", "FUENTE_Q0", "CONFIANZA_PROYECCION",
        "REDENCION", "ELASTICIDAD_FINAL", "UPLIFT", "Q0_DIA", "Q1_DIA",
        "GMV_BASE_DIA", "GMV_PROY_DIA", "GMV_INC_DIA", "UTIL_BASE_DIA", "UTIL_PROY_DIA", "UTIL_INC_DIA",
    ]
    out = df[cols_out].copy()
    out["DESC_EFECTIVO"] = (out["DESC_EFECTIVO"] * 100).round(1)
    out["DMAX_REAL"] = (out["DMAX_REAL"] * 100).round(1)
    out["UPLIFT"] = (out["UPLIFT"] * 100).round(0)
    out["ROTACION_REAL_RATIO"] = out["ROTACION_REAL_RATIO"].round(2)
    out["DIAS_ENTRE_VENTA_SKU_TIENDA"] = out["DIAS_ENTRE_VENTA_SKU_TIENDA"].round(1)
    out["DIAS_ENTRE_VENTA_CATEGORIA_BASELINE"] = out["DIAS_ENTRE_VENTA_CATEGORIA_BASELINE"].round(1)
    out = out.rename(
        columns={
            "DESC_EFECTIVO": "DESC_EFECTIVO_%",
            "MARGEN_OFERTA": "MARGEN_OFERTA_%",
            "UPLIFT": "UPLIFT_%",
            "DMAX_REAL": "DMAX_REAL_%",
        }
    )
    out = out.sort_values(["DIA_EJECUCION", "STORE_ID", "ESTRATEGIA"]).reset_index(drop=True)

    out_oferta = out[out["ESTRATEGIA"] != "Sin oferta"].reset_index(drop=True)
    out_descartados = out[out["ESTRATEGIA"] == "Sin oferta"].reset_index(drop=True)

    resumen = pd.DataFrame(
        {
            "Indicador": [
                "Fin de semana objetivo", "Universo (tienda x SKU)", "Ofertas asignadas", "% cobertura",
                "Descuento exhibido promedio", "Margen minimo en promocion",
                "Requieren aprobacion Comercial (d_max > 30%)", "Ofertas Tema Mundial",
                "Unidades base -> proyectadas (escenario base)",
                "GMV base -> proyectado (escenario base)", "Utilidad base -> proyectada (escenario base)",
                "Utilidad incremental total (base)", "Backtesting (n, MAE, sesgo)", "Guardrails aplicados",
            ],
            "Valor": [
                f"{weekend_inicio.date()} a {weekend_fin.date()}", len(df), len(ofer),
                f"{len(ofer)/len(df)*100:.1f}%", f"{ofer['DESC_EFECTIVO'].mean()*100:.1f}%",
                f"{ofer['MARGEN_OFERTA'].min():.2f}%", int(df["REQUIERE_APROBACION"].sum()),
                int(ofer["TEMA_MUNDIAL"].sum()),
                f"{ofer['Q0_DIA'].sum():,.0f} -> {ofer['Q1_DIA'].sum():,.0f}",
                f"${ofer['GMV_BASE_DIA'].sum():,.0f} -> ${ofer['GMV_PROY_DIA'].sum():,.0f}",
                f"${ofer['UTIL_BASE_DIA'].sum():,.0f} -> ${ofer['UTIL_PROY_DIA'].sum():,.0f}",
                f"${ofer['UTIL_INC_DIA'].sum():,.0f}",
                f"n={len(backtest)}, MAE={mae:.1f}pts, sesgo={bias:+.1f}pts" if len(backtest) else "sin datos",
                "Piso 22% por unidad + utilidad incremental >= 0 + tope autonomo 30%",
            ],
        }
    )
    por_dia = (
        ofer.groupby("DIA_EJECUCION")
        .agg(Ofertas=("SKU", "count"), GMV_incremental=("GMV_INC_DIA", "sum"), Utilidad_incremental=("UTIL_INC_DIA", "sum"))
        .round(0)
        .reset_index()
    )
    mecs = df["MECANICA"].value_counts().rename_axis("Mecanica").reset_index(name="SKU_tienda")

    leyenda = pd.DataFrame(
        [
            ("MOTIVO_SIN_OFERTA", "Por que la fila no lleva oferta (solo aplica en la hoja Descartados)"),
            ("REQUIERE_APROBACION / DMAX_REAL_%", "True si el colchon de margen real supera el tope autonomo de 30%; DMAX_REAL_% es ese colchon completo"),
            ("TEMA_MUNDIAL", "True si el nombre coincide con el patron de consumo 'ver el partido'"),
            ("ROTACION_REAL_RATIO", "Unidades/dia real del SKU dividido entre el promedio real de su categoria (rotacion.py) - 1.0 = en linea con su categoria, ya no depende solo del TAG_ROTACION estatico"),
            ("DIAS_ENTRE_VENTA_SKU_TIENDA / _CATEGORIA_BASELINE", "Dias promedio entre una venta y otra: del SKU, y el tipico de su categoria - para comparar ritmos de compra justo (ej. platano vs cepillo de dientes)"),
            ("FUENTE_BASELINE_CATEGORIA", "Categoria x tienda / Categoria (ambas tiendas) / Sin baseline - de donde salio el baseline de rotacion de la categoria"),
            ("ESTRATEGIA / DIA_EJECUCION", "Mecanica asignada y dia(s) del FDS en que corre"),
            ("PRECIO_OFERTA", "Precio por unidad cuando la mecanica se activa"),
            ("DESC_EFECTIVO_% / MARGEN_OFERTA_%", "Profundidad exhibida y pure margin de las unidades en promocion"),
            ("FUENTE_ELASTICIDAD / FUENTE_Q0", "SKU real / Categoria real / Supuesto declarado - de donde salio cada dato"),
            ("CONFIANZA_PROYECCION", "Alta: elasticidad propia + historial de promo real. Media: elasticidad propia, sin historial. Baja: fallback"),
            ("REDENCION", "Fraccion supuesta de unidades que se venden con la mecanica activa"),
            ("ELASTICIDAD_FINAL / UPLIFT_%", "Elasticidad usada (real o fallback) y crecimiento proyectado de unidades"),
            ("Q0_DIA / Q1_DIA", "Unidades/dia sin y con promocion"),
            ("GMV_*_DIA", "Venta en pesos del dia de oferta: base, proyectada, incremental"),
            ("UTIL_*_DIA", "Pure margin en pesos del dia de oferta: base, proyectada, incremental"),
        ],
        columns=["Columna", "Descripcion"],
    )

    with pd.ExcelWriter(ruta_salida, engine="openpyxl") as w:
        out_oferta.to_excel(w, sheet_name="Estrategia FDS", index=False)
        out_descartados.to_excel(w, sheet_name="Descartados", index=False)
        resumen.to_excel(w, sheet_name="Resumen", index=False, startrow=0)
        por_dia.to_excel(w, sheet_name="Resumen", index=False, startrow=len(resumen) + 3)
        mecs.to_excel(w, sheet_name="Resumen", index=False, startrow=len(resumen) + 3 + len(por_dia) + 3)
        escenarios_df.to_excel(w, sheet_name="Escenarios", index=False)
        backtest.to_excel(w, sheet_name="Backtesting", index=False)
        leyenda.to_excel(w, sheet_name="Leyenda", index=False)

        wb = w.book
        hdr = PatternFill("solid", start_color="1F4E79")
        verde, dorado, azul = (
            PatternFill("solid", start_color="E2EFDA"),
            PatternFill("solid", start_color="FFF2CC"),
            PatternFill("solid", start_color="DDEBF7"),
        )

        for hoja_nombre in ["Estrategia FDS", "Descartados"]:
            ws = wb[hoja_nombre]
            for c in ws[1]:
                c.font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
                c.fill = hdr
                c.alignment = Alignment(horizontal="center", vertical="center")
            ws.freeze_panes = "D2"
            ws.auto_filter.ref = ws.dimensions
            headers = [c.value for c in ws[1]]
            i_mun = headers.index("TEMA_MUNDIAL")
            i_apr = headers.index("REQUIERE_APROBACION")
            for row in ws.iter_rows(min_row=2):
                if hoja_nombre == "Estrategia FDS":
                    f = azul if row[i_apr].value else (dorado if row[i_mun].value else verde)
                    for cell in row:
                        cell.fill = f
            for i in range(1, ws.max_column + 1):
                ws.column_dimensions[get_column_letter(i)].width = {3: 50}.get(i, 14)

        for hoja in ["Resumen", "Escenarios", "Backtesting", "Leyenda"]:
            ws2 = wb[hoja]
            for c in ws2[1]:
                c.font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
                c.fill = hdr
            ws2.column_dimensions["A"].width = 46
            ws2.column_dimensions["B"].width = 56

    return out_oferta, out_descartados


def _slug_fecha(ts: pd.Timestamp) -> str:
    return ts.strftime("%b%d").lower()  # ej. jul03


def generar_ruta_salida(
    weekend_inicio: pd.Timestamp,
    weekend_fin: pd.Timestamp,
    carpeta: str = "data/output",
    run_ts: pd.Timestamp = None,
) -> str:
    """Nombre de archivo por corrida: `estrategia_fds_{finde}_{fecha_
    corrida}.xlsx`. Distinto fin de semana -> distinto archivo (nunca pisa
    una promo pasada). Misma corrida el mismo dia -> mismo nombre, se
    sobreescribe (es solo la ultima version de esa corrida, no historial)."""
    if run_ts is None:
        run_ts = pd.Timestamp.now()
    nombre = (
        f"estrategia_fds_{_slug_fecha(weekend_inicio)}_{_slug_fecha(weekend_fin)}"
        f"_{run_ts:%Y%m%d}.xlsx"
    )
    return str(Path(carpeta) / nombre)


def buscar_salida_historica(
    weekend_inicio: pd.Timestamp,
    weekend_fin: pd.Timestamp,
    carpeta: str = "data/output",
) -> str:
    """Encuentra el Excel ya generado para un fin de semana especifico (el
    mas reciente si hubo varias corridas), para post-mortem de una promo
    que ya paso. Si no hay ninguno con el nombre nuevo (con fechas), cae al
    nombre fijo `estrategia_fds.xlsx` por compatibilidad con corridas
    anteriores a esta convencion de nombres."""
    patron = f"estrategia_fds_{_slug_fecha(weekend_inicio)}_{_slug_fecha(weekend_fin)}_*.xlsx"
    candidatos = sorted(Path(carpeta).glob(patron))
    if candidatos:
        return str(candidatos[-1])
    legado = Path(carpeta) / "estrategia_fds.xlsx"
    if legado.exists():
        return str(legado)
    raise FileNotFoundError(
        f"No se encontro ningun Excel para el fin de semana "
        f"{weekend_inicio.date()} - {weekend_fin.date()} en {carpeta}"
    )


def construir_estrategia(
    cur,
    ruta_oportunidad: str,
    ruta_descuentos: str,
    ruta_salida: str,
    weekend_inicio: pd.Timestamp,
    weekend_fin: pd.Timestamp,
) -> pd.DataFrame:
    """Orquestador de punta a punta: arma la estrategia FDS y exporta el
    Excel final. `cur` es un cursor ya conectado a Snowflake (SSO desde el
    notebook, o la conexion que corresponda). Devuelve el `df` completo
    (util para inspeccionar en el notebook despues de correr esto)."""
    df = cargar_oportunidad(ruta_oportunidad)
    print(f"Universo (tienda x SKU): {len(df)}")

    df = excluir_comercial(df, ruta_descuentos, weekend_inicio, weekend_fin)
    df = agregar_catalogo_real(df, cur)
    df = agregar_rotacion_real(df, cur)
    df = calcular_margen_neto(df)
    df = marcar_elegibilidad(df)
    df = detectar_tema_mundial(df)
    df = calcular_dmax(df)

    athenea = elasticity.get_athenea(cur)
    df = df.merge(athenea, on=["SKU", "STORE_ID"], how="left")
    unidades_dia_real = elasticity.get_unidades_dia_fulfillment(cur)
    df = df.merge(unidades_dia_real, on=["SKU", "STORE_ID"], how="left")
    df = elasticity.aplicar_fallback_cascada(df)

    df = calcular_score_y_tope(df)
    df = asignar_mecanica(df)
    df = validar_piso(df)
    df = calcular_modelo_economico(df)
    df = aplicar_guardrail_economico(df)

    skus_elegibles = set(df.loc[df["ELEGIBLE"], "SKU"])
    promos_relevantes, skus_con_historial_promo = get_promo_historicas(cur, skus_elegibles)
    df = calcular_confianza(df, skus_con_historial_promo)
    ventas_historicas = get_ventas_historicas(cur, skus_elegibles)
    backtest, mae, bias = calcular_backtesting(df, promos_relevantes, ventas_historicas)

    df = asignar_dia_ejecucion(df)
    escenarios_df = calcular_escenarios(df)

    exportar_excel(df, backtest, escenarios_df, ruta_salida, weekend_inicio, weekend_fin, mae, bias)

    ofer = df[df["MECANICA"] != "Sin oferta"]
    print(f"Ofertas: {len(ofer)} de {len(df)} ({len(ofer)/len(df)*100:.1f}%)")
    if len(backtest):
        print(f"Backtesting: n={len(backtest)} | MAE={mae:.1f}pts | sesgo={bias:+.1f}pts")
    return df
